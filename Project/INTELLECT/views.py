from django.http import HttpResponse
#Обновил Никита, работает

import os
import openpyxl
import models
import repository
import utils
from order_804 import equipment_order_804


def parse_view(request):
    # Очистка старых данных для провайдера INTELLECT
    repository.ItemsRepository.delete_provider('INTELLECT')

    file_path = utils.get_xlsx_file(os.getcwd() + '/INTELLECT')

    if not file_path:
        print("No Excel file found in the INTELLECT directory.")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet is None:
            break

        image_loader = utils.SheetImageLoader(sheet)  # Initialize image loader for the sheet

        try:
            for row in sheet.iter_rows(min_row=2):
                num = row[0].value
                if num is None or num == "№":
                    continue

                size_color = row[3].value  # D //Размеры/цвет
                if size_color:
                    size_color = size_color.replace("(Ш×Г×В)", "").replace("\n", "").replace(" ", "")  # Убираем все лишнее из строки размеров

                item = models.Item(
                    num=num,
                    name_provider=row[1].value,
                    provider='INTELLECT',
                    name_804=equipment_order_804.get(num),
                    article=row[12].value,
                    size=size_color,
                    cost=row[8].value,
                    img=utils.get_value_from_merged_image(sheet, image_loader, row[4]),
                )

                print(item.name_provider)
                repository.ItemsRepository.create(item)

        except Exception as e:
            print(f"An error occurred while processing sheet {sheet_name}: {e}")
            continue

    return HttpResponse('OK')

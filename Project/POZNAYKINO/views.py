#Обновил, работает
from django.http import HttpResponse
import os
import openpyxl
import models
import utils
import re
from order_804 import equipment_order_804
import repository


def remove_non_numeric_chars(input_string):
    return re.sub(r'[^0-9.]', '', input_string)


def parse_view(request):
    # Очистка старых данных для провайдера POZNAYKINO
    repository.ItemsRepository.delete_provider('POZNAYKINO')

    file_path = utils.get_xlsx_file(os.getcwd() + '/POZNAYKINO')
    sheets_dict = {

        # 0 // num/
        # 1 name804/
        # 2 name_provider/
        # 3 описание/
        # 4 ед измерения/
        # 5 цена/
        # 6 артикул/
        # 7 изображение
        "УОК и АПК": [6, 6, 1, 2, 3, 4, 0],
        "Цифровые лаб": [6, 6, 1, 2, 3, 4, 0],
        "Солнечный город": [6, 6, 1, 2, 3, 4, 0, 1],
        "Умная ферма.Теплица": [6, 6, 1, 2, 3, 4, 0, 1],
        "потолочная система ": [5, 5, 1, 2, 3, 4, 0, 1],
        "Логопедическое оборудование": [6, 6, 1, 2, 3, 4, 0],
        "Роботы": [5, 5, 1, 2, 3, 4, 0],
        "Печатная продукция": [19, 19, 1, 3, 4, 6, 0],
        "Карты": [-1, -1, 1, 2, 3, 4, 0],

    }
    if not file_path:
        print("No Excel file found in the POZNAYKINO directory.")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet is None:
            break

        image_loader = utils.SheetImageLoader(sheet)  # Initialize image loader for the sheet

        try:
            for row in sheet.iter_rows(min_row=2):
                if sheets_dict[sheet_name][0] != -1:
                    num = row[sheets_dict[sheet_name][0]].value
                    if num is None:
                        num = "-"
                    num = remove_non_numeric_chars(num)
                else:
                    num = "-"

                if sheets_dict[sheet_name][1] != -1:
                    temp = row[sheets_dict[sheet_name][1]].value
                    name_804 = equipment_order_804.get(row[sheets_dict[sheet_name][1]].value)
                else:
                    name_804 = "-"

                name_provider = (f'{row[sheets_dict[sheet_name][2]].value}, '
                                 f'Характеристика: {row[sheets_dict[sheet_name][3]].value}, '
                                 f'Ед. измерения: {row[sheets_dict[sheet_name][4]].value}')

                provider = "POZNAYKINO"
                cost = row[sheets_dict[sheet_name][5]].value

                size = '-'
                article = row[sheets_dict[sheet_name][6]].value
                if article is None:
                    continue
                if "Артикул" in article:
                    continue

                image_data = "-"
                try:
                    if len(sheets_dict[sheet_name]) > 7:
                        cell_image = row[sheets_dict[sheet_name][7]]
                        if cell_image:
                            image_data = utils.get_value_from_merged_image(sheet, image_loader, cell_image)
                except Exception:
                    image_data = "-"

                item = models.Item(
                    num=num,
                    name_provider=name_provider,
                    provider=provider,
                    name_804=equipment_order_804.get(num),
                    cost=cost,
                    size=size,
                    article=article,
                    img=image_data
                )

                #print(item.name_804)
                repository.ItemsRepository.create(item)

        except Exception as e:
            print(f"An error occurred while processing sheet {sheet_name}: {e}")
            continue

    return HttpResponse('OK')

if __name__ == "__main__":
    parse_view()
